"""
Shared writer for the Initiatives Tracker file.

Writes a month X tracker that is FORMAT-IDENTICAL to the month X-1 tracker: it
opens the X-1 file itself as the template (so column widths, fonts, borders, row
heights and the status colour-coding are preserved verbatim), refreshes the title
month/year, overwrites the data cell values with month X's data, and re-applies
the status fill colour to each row according to its NEW status (so a row that
changed status gets the right colour, not the stale X-1 colour).

If no template is available it falls back to a plain (unformatted) workbook.

Both Step 1 (initial fill) and Step 2 (after Status change is computed) call this
so the file on disk always reflects the latest state.
"""
from copy import copy

import pandas as pd

from .. import config


def _clean(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str) and v.strip().lower() == "nan":
        return None
    return v


def _status_col_index(ws, header_row_1based: int):
    """1-based column index whose header (on the header row) is 'Status'."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=header_row_1based, column=c).value or "").strip().lower() == "status":
            return c
    return None


def _status_fill_map(ws, header_row_1based: int, status_col: int) -> dict:
    """Map each status value -> a copy of the fill used for it in the template."""
    out = {}
    if not status_col:
        return out
    for r in range(header_row_1based + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=status_col)
        val = str(cell.value or "").strip()
        if val and val not in out and cell.fill and cell.fill.patternType:
            out[val] = copy(cell.fill)
    return out


def _save(wb, out_path):
    """Save, falling back to the first writable '(n)' alternate if locked."""
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        for n in range(1, 100):
            alt = out_path.with_name(f"{out_path.stem} ({n}){out_path.suffix}")
            try:
                wb.save(alt)
                return alt
            except PermissionError:
                continue
        raise


def write_tracker(raw: pd.DataFrame, header_idx: int,
                  current: pd.DataFrame, month: int, year: int,
                  template_path=None):
    """Write month X tracker preserving the X-1 template; returns the output Path."""
    out_path = config.INITIATIVES_TRACKER_DIR / f"Initiatives tracker thang {month}-{year}.xlsx"
    config.INITIATIVES_TRACKER_DIR.mkdir(parents=True, exist_ok=True)

    title = (f"Initiatives Tracker - status as of end "
             f"{config.get_month_name(month)} {year}")
    n_cols = raw.shape[1]

    # ---- Preferred path: clone the X-1 file so formatting is identical ----
    if template_path is not None:
        try:
            from pathlib import Path
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            tpath = Path(template_path)
            if tpath.exists():
                wb = load_workbook(tpath)
                ws = wb.active
                ws.title = f"Initiatives tracker {month}-{year}"

                header_row = header_idx + 1               # 1-based header row
                status_col = _status_col_index(ws, header_row)
                fill_map = _status_fill_map(ws, header_row, status_col)
                no_fill = PatternFill(fill_type=None)

                ws.cell(row=1, column=1).value = title    # refresh title month/year

                first_data = header_idx + 2               # 1-based first data row
                for i, (_, row) in enumerate(current.iterrows()):
                    excel_row = first_data + i
                    for c in range(n_cols):
                        col_name = current.columns[c] if c < len(current.columns) else None
                        val = _clean(row[col_name]) if col_name is not None else None
                        ws.cell(row=excel_row, column=c + 1).value = val
                    # Re-colour the status cell for THIS month's status
                    if status_col:
                        sval = str(_clean(row[current.columns[status_col - 1]]) or "").strip() \
                            if status_col - 1 < len(current.columns) else ""
                        ws.cell(row=excel_row, column=status_col).fill = \
                            fill_map.get(sval, no_fill)

                # Drop any leftover template data rows beyond month X's data
                extra = first_data + len(current)
                if ws.max_row >= extra:
                    ws.delete_rows(extra, ws.max_row - extra + 1)

                return _save(wb, out_path)
        except Exception as e:
            print(f"[WARNING] tracker template clone failed ({e}); writing plain workbook")

    # ---- Fallback: plain workbook (values only, no formatting) ----
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Initiatives tracker {month}-{year}"
    for r in range(header_idx + 1):
        values = [_clean(v) for v in raw.iloc[r].tolist()]
        if r == 0:
            values[0] = title
        ws.append(values)
    for _, row in current.iterrows():
        values = [
            _clean(row[current.columns[c]]) if c < len(current.columns) else None
            for c in range(n_cols)
        ]
        ws.append(values)
    return _save(wb, out_path)
